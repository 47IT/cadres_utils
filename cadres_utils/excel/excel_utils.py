import io
import os.path
from copy import copy
from datetime import date

import pandas as pd
from pandas import DataFrame
from openpyxl import Workbook

from cadres_utils.file.utils import get_random_string


def save_default_excel_file(df: DataFrame, save_path: str, export_index=False, file_name: str = None) -> str:
    if file_name is None:
        file_hash = get_random_string(100)
        res_file_name = f'{file_hash}.xlsx'
    else:
        res_file_name = f'{file_name}.xlsx'

    file_path = os.path.join(save_path, res_file_name)
    writer = pd.ExcelWriter(
        file_path,
        engine="xlsxwriter",
        datetime_format="dd.mm.yyyy",
        date_format="dd.mm.yyyy",
    )
    df.to_excel(writer, index=export_index)
    writer.close()
    return file_path


def save_default_excel_to_io_stream(df: DataFrame, export_index=False) -> io.BytesIO:
    doc_io = io.BytesIO()
    writer = pd.ExcelWriter(
        doc_io,
        engine="xlsxwriter",
        datetime_format="dd.mm.yyyy",
        date_format="dd.mm.yyyy",
    )
    df.to_excel(writer, index=export_index)

    writer.close()
    doc_io.seek(0)
    return doc_io


def get_default_file_name(base_name: str, start_date: date, end_date: date, file_extension: str = '.xlsx') -> str:
    start_date = start_date.strftime('%Y-%m-%d')
    end_date = end_date.strftime('%Y-%m-%d')

    if start_date == end_date:
        tmp_str = start_date
    else:
        tmp_str = f'{start_date} - {end_date}'
    return f'{base_name} - {tmp_str}{file_extension}'


def save_workbook_to_file(wb: Workbook, save_path: str) -> str:
    file_hash = get_random_string(100)
    res_file_name = f'{file_hash}.xlsx'
    file_path = os.path.join(save_path, res_file_name)
    wb.save(file_path)
    return file_path


def work_book_2_io_stream(wb: Workbook) -> io.BytesIO:
    doc_io = io.BytesIO()
    wb.save(doc_io)
    doc_io.seek(0)

    return doc_io


def copy_row_styles_and_formulas(
        sheet, src_row_index: int, start_row_index: int, end_row_index: int, orig_template_formula_row: int | None = None
):
    last_col_index = sheet.max_column
    template_styles = __create_styles(sheet, src_row_index)
    for row in range(start_row_index, end_row_index + 1):
        for col in range(1, last_col_index + 1):
            new_cell = sheet.cell(row=row, column=col)
            if col in template_styles:
                curr_style = template_styles[col]
                new_cell.font = curr_style['font']
                new_cell.fill = curr_style['fill']
                new_cell.border = curr_style['border']
                new_cell.alignment = curr_style['alignment']
                new_cell.number_format = curr_style['number_format']

            # Copy formula, adjusting row references
            if orig_template_formula_row:
                ref_cell = sheet.cell(row=src_row_index, column=col)
                if ref_cell.data_type == 'f':
                    formula = ref_cell.value
                    # Adjust row references in the formula
                    new_formula = formula.replace(str(orig_template_formula_row), str(row))
                    new_cell.value = new_formula


def update_row_formulas(sheet, row_index: int, orig_template_row_shift: int):
    last_col_index = sheet.max_column
    for col in range(1, last_col_index + 1):
        cell = sheet.cell(row=row_index, column=col)
        if cell.data_type == 'f':
            formula = cell.value
            new_formula = formula.replace(str(orig_template_row_shift), str(row_index))
            cell.value = new_formula


def write_to_cell(write_sheet, write_row_num: int, write_col_num: int, value: str | int | float | None):
    if value is not None: # 0 value should be written
        new_cell = write_sheet.cell(row=write_row_num, column=write_col_num)
        new_cell.value = value


def __create_styles(sheet, template_row) -> dict:
    template_styles = {}
    max_col = sheet.max_column

    for col in range(1, max_col + 1):
        template_cell = sheet.cell(row=template_row, column=col)
        template_styles[col] = {
            'font': copy(template_cell.font),
            'fill': copy(template_cell.fill),
            'border': copy(template_cell.border),
            'alignment': copy(template_cell.alignment),
            'number_format': copy(template_cell.number_format)
        }

    return template_styles